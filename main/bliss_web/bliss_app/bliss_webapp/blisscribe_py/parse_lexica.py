# -*- coding: utf-8 -*-
"""
PARSE_LEXICA:

    Contains a class for parsing a language-to-Blissymbols
    dictionary from plaintext (.txt) and Excel (.xlsx) files.

    Allows ease of changing Bliss dictionaries, including
    adding languages beyond English.

    Throughout Blisscribe, "lemma" is meant to be the dictionary
    entry of a word, while "lexeme" is meant to be any form of a
    word.  All lemmas are lexemes, but only 1 in a set of lexemes
    is chosen as the lemma.
    e.g. dog == lemma (and == lexeme)
         dogs == lexeme (and != lemma)

    Alphabetical list of part-of-speech tags used in the
    Penn Treebank Project:

    Number  Tag     Description
    1.      CC      Coordinating conjunction
    2.      CD	    Cardinal number
    3.	    DT	    Determiner
    4.	    EX	    Existential there
    5.	    FW	    Foreign word
    6.	    IN	    Preposition or subordinating conjunction
    7.	    JJ	    Adjective
    8.	    JJR	    Adjective, comparative
    9.	    JJS	    Adjective, superlative
    10.	    LS	    List item marker
    11.	    MD	    Modal
    12.	    NN	    Noun, singular or mass
    13.	    NNS	    Noun, plural
    14.	    NNP	    Proper noun, singular
    15.	    NNPS	Proper noun, plural
    16.	    PDT	    Predeterminer
    17.	    POS	    Possessive ending
    18.	    PRP	    Personal pronoun
    19.	    PRP$	Possessive pronoun
    20.	    RB	    Adverb
    21.	    RBR	    Adverb, comparative
    22.	    RBS	    Adverb, superlative
    23.	    RP	    Particle
    24.	    SYM	    Symbol
    25.	    TO	    to
    26.	    UH	    Interjection
    27.	    VB	    Verb, base form
    28.	    VBD	    Verb, past tense
    29.	    VBG	    Verb, gerund or present participle
    30.	    VBN	    Verb, past participle
    31.	    VBP	    Verb, non-3rd person singular present
    32.	    VBZ	    Verb, 3rd person singular present
    33.	    WDT	    Wh-determiner
    34.	    WP	    Wh-pronoun
    35.	    WP$	    Possessive wh-pronoun
    36.	    WRB	    Wh-adverb
"""
import os
from openpyxl import load_workbook
from PIL import Image
import re
import blissymbols
from blissymbols import Blissymbol, BLISS_TO_UNICODE, UNICODE_TO_BLISS, NEW_BLISSYMBOLS

LAST_BLISS_ENCODING = blissymbols.LAST_BLISS_ENCODING

BLISS_SUPPORTED_LANGUAGES = set(["English",
                                 "Swedish",
                                 "Norwegian",
                                 "Finnish",
                                 "Hungarian",
                                 "German",
                                 "Dutch",
                                 "Afrikaans",
                                 "Russian",
                                 "Latvian",
                                 "Polish",
                                 "French",
                                 "Spanish",
                                 "Portuguese",
                                 "Italian",
                                 "Danish"])


class LexiconParser:
    FILE_PATH = os.path.dirname(os.path.realpath(__file__))
    IMG_PATH = FILE_PATH + blissymbols.IMG_LOCN
    LEXICA_PATH = FILE_PATH + "/resources/lexica/"
    LEXICON_PATH = LEXICA_PATH + "universal bliss lexicon.xlsx"
    LEXICON_COLS = ["BCI-AV#",
                    "English",
                    "POS",
                    "Derivation - explanation",
                    "BCI-AV#",
                    "Swedish",
                    "BCI-AV#",
                    "Norwegian",
                    "BCI-AV#",
                    "Finnish",
                    "BCI-AV#",
                    "Hungarian",
                    "BCI-AV#",
                    "German",
                    "BCI-AV#",
                    "Dutch",
                    "BCI-AV#",
                    "Afrikaans",
                    "BCI-AV#",
                    "Russian",
                    "BCI-AV#",
                    "Latvian",
                    "BCI-AV#",
                    "Polish",
                    "BCI-AV#",
                    "French",
                    "BCI-AV#",
                    "Spanish",
                    "BCI-AV#",
                    "Portuguese",
                    "BCI-AV#",
                    "Italian",
                    "BCI-AV#",
                    "Danish"]

    def __init__(self, translator):
        self.translator = translator

    def parseLexicon(self, filename):
        """
        Parses plaintext file with given filename.
        Returns a dict with all words in lexicon
        as keys and corresponding lemma forms as values.
        ~
        Each new lexical entry should be separated by "\n",
        while inflected forms should be separated by ",".
        ~
        Assumes first word on every line is the lemma form
        of all subsequent words on the same line.
        ~
        N.B. The same lemma value will often belong to multiple keys.

        e.g. "kota, kocie, kot" -> {"kota":"kota", "kocie":"kota", "kot":"kota"}

        :param filename: str, filename of .txt file for lexicon
        :return: dict, where...
            key (str) - inflected form of a word
            val (str) - lemma form of inflected word
        """
        lexicon = {}

        with open(self.FILE_PATH + filename, "rb") as lex:
            for entry in lex:
                entry = entry.decode("utf-8")
                entry = entry.strip("\n")
                entry = entry.strip("\r")
                inflexions = entry.split(",")
                lemma = inflexions[0]

                for inflexion in inflexions:
                    lexicon[inflexion.strip()] = lemma

        return lexicon

    def parseFrenchLexicon(self, filename):
        """
        Parses plaintext file for French lexicon with
        given filename.  Returns a dict with all lexemes
        in French lexicon as keys and corresponding
        lemma forms as values.
        ~
        Each new lexical entry should be separated by "\n",
        while inflected forms should be separated by ",".
        ~
        Assumes second word on every line is the lemma form
        of previous words on the same line.
        ~
        N.B. The same lemma value will often belong to
        multiple lexemes.

        e.g. "grand, grande, grands" -> {"grand":"grand", "grande":"grand", "grands":"grand"}

        :param filename: str, filename of .txt file for French lexicon
        :return: dict, where...
            key (str) - any lexical form of a word
            val (str) - lemmatized form of lexeme
        """
        lexicon = {}

        with open(self.FILE_PATH + filename, "rb") as lex:
            for line in lex:
                line = line.decode("utf-8")
                line = line.strip("\n")
                if line[-1] == "=":
                    lemma = line[:-2]
                    lexicon[lemma] = lemma
                else:
                    entry = line.split("\t")
                    lemma = entry[1]
                    lexeme = entry[0]
                    lexicon[lexeme] = lemma

        return lexicon

    def getTabFile(self, lang_code):
        """
        Retrieves a multilingual WordNet tab file for the
        given language.
        ~
        If no such tab file exists, returns None.

        :param lang_code: str, 3-character ISO language code
        :return: tab file, WordNet file for given language
        """
        try:
            return open(self.FILE_PATH + "/resources/omw_tabs/" +
                        "wn-cldr-" + lang_code + ".tab")
        except Exception:
            return None

    # ATOMIC BLISSYMBOLS
    # ==================
    def getBlissEncoding(self, filename=None):
        """
        Reads plaintext file with given filename and returns
        a list of Blissymbol names.
        ~
        Output conforms to encoding suggested here:
        http://std.dkuug.dk/JTC1/SC2/WG2/docs/n1866.pdf

        :param filename: str, .txt file with Blissymbol names
        :return: List[str], Blissymbol names
        """
        if filename is None:
            filename = self.LEXICA_PATH + "blissymbol encoding.txt"
        else:
            filename = self.FILE_PATH + filename

        defns = []

        with open(filename, "rb") as bliss_in:
            for bliss_name in bliss_in:
                name = bliss_name[11:-1].lower()
                defns.append(name)

        return defns

    def writeBlissEncoding(self, unicode_keys=False):
        """
        Writes and returns a dict linking Blissymbol names to unicode values,
        or, if unicode_keys=True, linking unicode keys to Blissymbol names.
        ~
        Output conforms to encoding suggested here:
        http://std.dkuug.dk/JTC1/SC2/WG2/docs/n1866.pdf
        ~
        Used to update BLISS_TO_UNICODE and UNICODE_TO_BLISS as needed.

        :param unicode_keys: bool, whether output dict keys are unicode
        :return: dict, where...
            key (str) - word definition (unicode if unicode_keys=True)
            val (str) - unicode (word definition if unicode_keys=True)
        """
        defns = self.getBlissEncoding()
        encodings = {}

        if unicode_keys:
            title = "UNICODE_TO_BLISS"
        else:
            title = "BLISS_TO_UNICODE"

        write_filename = self.LEXICA_PATH + "bliss_encoding.py"
        idx = "3200"

        with open(write_filename, "a") as bliss_out:
            bliss_out.write(title + " = {")

            for defn in defns:
                uni = "U+" + idx  # make unicode key

                if unicode_keys:
                    key = uni
                    val = defn
                else:
                    key = defn
                    val = uni

                encoding = '\n    "' + key + '": ["' + val + '"],'
                bliss_out.write(encoding)
                encodings[key] = [val]

                idx = int(idx, 16) + 1
                idx = hex(idx)[2:]

            bliss_out.write("\n    }\n\n")

        global LAST_BLISS_ENCODING
        LAST_BLISS_ENCODING = idx

        return encodings

    def getBlissAtoms(self, filename=None):
        """
        Reads list of Blissymbol atoms from plaintext file with given
        filename and returns a list of Blissymbol atoms.

        :param filename: str, .txt file with Blissymbol names
        :return: list, of Blissymbol atom names (in English)
        """
        bliss_atoms = {}
        if filename is None:
            filename = self.LEXICA_PATH + "blissymbol encoding.txt"
        else:
            filename = self.LEXICA_PATH + filename

        with open(filename, "rb") as blissymbols:

            for blissymbol in blissymbols:
                name = blissymbol[11:-1].lower()
                unicode = BLISS_TO_UNICODE[name]
                name = name.replace(" ", "_")
                bliss_atoms[name] = unicode

        return bliss_atoms

    def showBlissAtom(self, filename):
        img = Image.open(self.IMG_PATH+filename)
        img.show()

    def showBlissAtoms(self, bliss_atoms):
        for atom in bliss_atoms:
            self.showBlissAtom(bliss_atoms[atom])

    # ILI FUNCTIONS
    # =============
    def readILIMapping(self):
        """
        Reads plaintext file with mapping from Princeton WordNet
        to ILI (Interlingual Language Index), a conceptual dictionary.
        ~
        Used for cross-lingual translation.

        :return: List[ILIEntry], list of ILI definitions
        """
        defns = []
        with open(self.LEXICA_PATH + "ili-wn-mapping.txt", "r") as ili:
            lines = ili.readlines()[10:]
            for defn in lines:
                defns.append(self.cleanILIDefn(defn))
        return defns

    def writeILIMapping(self, clean_defns):
        """
        Writes to plaintext file with mapping from Princeton WordNet
        to ILI (Interlingual Language Index), a conceptual dictionary.
        ~
        Used for cross-lingual translation.

        :param clean_defns: List[ILIEntry], list of word entries to write
        :return: None
        """
        out = self.LEXICA_PATH + "ili-wn-mapping-cleaned.txt"
        open(out, 'w').close()  # wipe file before writing

        with open(out, "a") as ili:
            for defn in clean_defns:
                ili.write(str(defn) + "\n")

    def readWriteILIMapping(self):
        """
        Reads ILI mapping and writes cleaned definitions
        to file.

        :return: None
        """
        clean_defns = self.readILIMapping()
        self.writeILIMapping(clean_defns)

    def writeBlissWordnet(self, blissnet):
        """
        Writes mapping between Blissymbol unicodes and
        Synsets to plaintext file.

        :param bliss_dict: dict, where...
            key (str) - Blissymbol's unicode representation
            val (List[Synset]) - synsets associated with Blissymbol
        :return: None
        """
        out = self.LEXICA_PATH + "blissnet.txt"
        open(out, 'w').close()  # wipe file before writing

        with open(out, "a") as wordnet:
            wordnet.write(self.dictToStr(blissnet))

    def blissDictToWordnet(self, bliss_dict):
        """
        Returns a dictionary of Bliss unicodes and synsets
        from the given bliss_dict.
        ~
        N.B. Output will be most accurate with a multilingual
        Bliss dictionary.

        :param bliss_dict: dict, where...
            key (str) - Blissymbol word in English
            val (List[Blissymbol]) - corresponding Blissymbols with
                translations in all languages
        :return: dict, where...
            key (str) - Blissymbol's unicode representation
            val (List[Synset]) - synsets associated with Blissymbol
        """
        wordnet = {}
        for word in bliss_dict:
            for blissymbol in bliss_dict[word]:
                uni = blissymbol.getUnicode()
                synsets = blissymbol.getSynsets()
                wordnet[uni] = synsets
        return wordnet

    # GETTERS
    # =======
    def getImgFilenames(self):
        """
        Reads the given XLS file for a cross-lingual Bliss
        dictionary and returns a list of image filenames.

        :param filename: str, name of XLS file
        :return: List[str], image filenames
        """
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        img_col = 2

        for col in sheet.iter_cols(min_col=img_col, max_col=img_col):
            imgs = [str(row.value) + ".png" for row in col[1:]]
            return imgs

    def makeBlissymbol(self):
        """
        Prompts user for new Blissymbol entry information.

        :return: Blissymbol, represents 1 Bliss lexical entry
        """
        bliss_name = raw_input("What do you call your new Blissymbol? ")
        bliss_filename = bliss_name + ".png"
        print("Which part of speech is this? ")
        code = Blissymbol.POS_COLOUR_CODE

        for pos in code:
            print(str(pos) + ":\t" + str(code[pos]))

        pos = raw_input("")
        print("Which atomic Blissymbols is this made of? ")
        derivations = raw_input("Separate them by commas: ")
        derivations = derivations.split(",")
        derivs = ""
        derivs += "("

        for idx in range(0, len(derivations)):
            derivs += derivations[idx]
            if idx == len(derivations)-1:
                derivs += ")"
            else:
                derivs += " + "

        translations = {}

        for language in BLISS_SUPPORTED_LANGUAGES:
            print("What is/are the translation(s) in " + language + "? ")

            try:
                translation = raw_input("Separate by commas if necessary: ")
            except SyntaxError:
                continue
            else:
                translation = translation

            translations.setdefault(language, [])
            translations[language].append(translation)

        blissymbol = Blissymbol(img_filename=bliss_filename, pos=pos, derivation=derivs,
                                translations=translations, translator=self.translator)
        return blissymbol

    def makeBlissXLXSEntry(self, blissymbol):
        """
        Converts the given Blissymbol to a list of
        information constituting a Bliss lexicon entry.
        ~
        Must be in order of LEXICON_COLS.

        :param blissymbol: Blissymbol, symbol to convert to entry
        :return: List[str], Bliss lexicon entry
        """
        row = []
        bci_col = self.LEXICON_COLS[0]
        pos_col = self.LEXICON_COLS[2]
        deriv_col = self.LEXICON_COLS[3]
        unicode = blissymbol.getUnicode() #[0]

        for col in self.LEXICON_COLS:
            if col == bci_col:
                row.append("C+" + unicode[2:])
                continue
            elif col == pos_col:
                row.append(blissymbol.getPosCode())
            elif col == deriv_col:
                row.append(blissymbol.getDerivation())
            else:
                translations = blissymbol.getTranslation(col)
                #translations = ",".join(translations)
                row.append(translations)

        return row

    def addBlissEntry(self, blissymbol=None):
        """
        Adds data to bliss lexicon file for later use.
        ~
        If no Blissymbol is provided, prompts you to create your
        own custom symbol.

        :param blissymbol: Blissymbol, entry to add to Bliss lexicon
        :return: None
        """
        if not blissymbol:
            blissymbol = self.makeBlissymbol()

        bliss_entry = self.makeBlissXLXSEntry(blissymbol)

        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        sheet.append(bliss_entry)

        book.save(self.LEXICON_PATH)

    def extendBlissEntry(self, blissymbol):
        """
        Extends data in bliss lexicon file for later use.

        :param bliss_entry: Blissymbol, entry to add to Bliss lexicon
        :return: None
        """
        print("extending entry for ", blissymbol.getBlissName())
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        start_row = sheet.min_row
        end_row = sheet.max_row

        for row_idx in range(start_row, end_row):
            cell = sheet.cell(row=row_idx, column=2)
            eng_word = cell.value
            eng_words = eng_word.split(",")
            for synonym in eng_words:
                synonym = synonym.strip()
                if synonym == blissymbol.getBlissName():
                    print("found English Blissymbol synonym: ", eng_word)
                    translations = blissymbol.getTranslations()
                    for language in translations:
                        defns = translations[language]
                        if len(defns) != 0:
                            lang_idx = self.LEXICON_COLS.index(language)
                            edit = sheet.cell(row=row_idx, column=lang_idx+1)
                            if type(edit.value) != str:
                                edit.value = str(edit.value)
                            for defn in defns:
                                if defn not in edit.value.split(","):
                                    edit.value += ","
                                    edit.value += defn
                            print("edited ", language)
                    break

        book.save(self.LEXICON_PATH)

    def getDefns(self, language):
        """
        Reads the given XLSX file for a cross-lingual Bliss
        dictionary and returns a list of the given language's
        words shared by the Bliss lexicon.

        :param language: str, output list's desired language
        :return: List[str], given language's Bliss words
        """
        assert language in BLISS_SUPPORTED_LANGUAGES

        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]

        defns = []

        for col in sheet.iter_cols():
            if col[0].value == language:
                defns = [row.value for row in col[1:]]
                break

        if len(defns) == 0:
            raise IOError(language + " is not in Blissymbolics lexicon.")

        return defns

    def getPartsOfSpeech(self):
        """
        Reads the given XLS file for a cross-lingual Bliss
        dictionary and returns an ordered list of parts of speech.

        :param filename: str, name of XLS file
        :return: List[str], list of parts of speech
        """
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        pos_col = 3

        for col in sheet.iter_cols(min_col=pos_col, max_col=pos_col):
            pos = [row.value for row in col[1:]]
            return pos

    def getDerivations(self):
        """
        Reads the given XLS file for a cross-lingual Bliss
        dictionary and returns an ordered list of derivations.

        :param filename: str, name of XLS file
        :return: List[str], list of derivations
        """
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        deriv_col = 4

        for col in sheet.iter_cols(min_col=deriv_col, max_col=deriv_col):
            derivations = [unicode.encode(row.value, 'ascii', errors='ignore') for row in col[1:]]
            return derivations

    def getDefnWords(self, defn):
        """
        Separates given defn by commas and returns the resulting list.

        :param defn: str, Blissymbol definition
        :return: List[str],
        """
        words = []
        if defn is not None:
            if not self.isOld(defn):  # no outdated definitions
                for word in defn.split(","):
                    word = self.parseAlphabetic(word)
                    if not self.isIndicator(word):
                        word = self.stripParens(word)
                    if len(word) != 0:
                        words.append(word)
        return words

    def isIndicator(self, word):
        """
        Returns True if this Blissymbol word is an indicator.

        :param word: str, Blissymbol word from bliss_dict
        :return: bool, True if given word is an indicator
        """
        return word[:9] == "indicator"

    def isOld(self, word):
        """
        Returns True if this Blissymbol word is marked OLD.

        :param word: str, Blissymbol word from bliss_dict
        :return: bool, True if given word is old
        """
        return word[-5:] == "(OLD)"

    def isLetter(self, img_fn):
        """
        Returns True if given image filename suggests that its
        corresponding Blissymbol is a letter of the alphabet.
        ~
        Checks if img_fn ends with "ercase" (for uppercase or lowercase).
        ~
        Used for excluding alphabetic-only Bliss characters.

        :param img_fn: str, image filename of a Blissymbol
        :return: bool, whether given image is an alphabetic letter
        """
        return img_fn[-11:] == "ercase).png"

    def getBlissDict(self, language):
        """
        Creates 2 dicts with words in chosen language and English
        as keys and Blissymbol images as vals.  Returns a tuple of
        both dicts.
        ~
        If input language is English, returns an English dict and
        an empty dict.  Otherwise, returns the given language's
        dict first in the tuple and the English dict second.
        ~
        If a definition contains multiple words, this function
        splits the definition and adds each word to the dict as
        separate entries linking to the same Blissymbol.
        ~
        If a word in the file has no corresponding Blissymbol,
        the {key,val} pair is not added to the output dict.

        :param language: str, first dictionary's language
        :return: tuple(dict, dict) in language & English, where in each dict...
            key (str) - words in specified language
            val (List[Blissymbol]) - corresponding Blissymbols
        """
        # TODO: parse parenthetical endings more gracefully
        lang_bliss_dict = {}
        defns = self.getDefns(language)
        imgs = self.getImgFilenames()
        parts_of_speech = self.getPartsOfSpeech()
        derivations = self.getDerivations()

        if language != "English":
            eng_bliss_dict = {}
            eng_defns = self.getDefns("English")
        else:
            eng_bliss_dict = lang_bliss_dict
            eng_defns = defns

        idx = 0

        for eng_defn in eng_defns:  # Bliss translations are in English
            if eng_defn is None or self.isLetter(imgs[idx]):  # don't translate incoherent symbols
                idx += 1
                continue

            defn = defns[idx]
            words = self.getDefnWords(defn)  # synonyms
            defn = u"" if defn is None else defn
            eng_words = self.getDefnWords(eng_defn)

            print(u"creating defn-img dict with " + (u"" if defn is None else defn) + u" / " + eng_defn)

            if len(eng_words) != 0:  # don't translate nonexistent words
                translations = {}
                translations["English"] = eng_words
                if len(words) != 0:
                    translations[language] = words

                bliss_word = Blissymbol(img_filename=imgs[idx],
                                        pos=parts_of_speech[idx],
                                        derivation=derivations[idx],
                                        #language=language,
                                        translations=translations,
                                        translator=self.translator)

                translations = bliss_word.getTranslations()

                for lang in translations:
                    if language != "English" and lang == "English":
                        bliss_dict = eng_bliss_dict
                    else:
                        bliss_dict = lang_bliss_dict

                    for translation in translations[lang]:
                        bliss_dict.setdefault(translation, [])
                        bliss_dict[translation].append(bliss_word)
                        if lang == "English":
                            bliss_word.addBlissAndUnicode(translation)

            print("\n")
            idx += 1

        return (lang_bliss_dict, eng_bliss_dict)

    def getMultilingualBlissDict(self):
        """
        Creates and returns a multilingual Bliss dictionary with
        English keys and Blissymbol values.  Each created
        Blissymbol is initialized with translations in every
        supported language.
        ~
        If a definition contains multiple words, this function
        splits the definition and adds each word to the dict as
        separate entries linking to the same Blissymbol.
        ~
        If a word in the file has no corresponding Blissymbol,
        the {key,val} pair is not added to the output dict.
        ~
        Used to achieve the most accurate synset estimations by
        cross-referencing all language's synsets.

        :param language: str, dictionary keys' language
        :return: dict, where...
            key (str) - words in English
            val (List[Blissymbol]) - corresponding Blissymbols with
                translations in all languages
        """
        lang_bliss_dict = {}
        defns = self.getDefns("English")
        #wn_langs = self.translator.WORDNET_LANGS
        #print(wn_langs)
        langs = BLISS_SUPPORTED_LANGUAGES #.intersection(wn_langs)
        multi_defns = {lang: self.getDefns(lang)
                       for lang in langs}
        imgs = self.getImgFilenames()
        parts_of_speech = self.getPartsOfSpeech()
        derivations = self.getDerivations()

        idx = 0

        for eng_defn in defns:  # Bliss translations are in English
            if eng_defn is None or self.isLetter(imgs[idx]):  # don't translate incoherent symbols
                idx += 1
                continue

            defn = defns[idx]
            defn = u"" if defn is None else defn
            eng_words = self.getDefnWords(eng_defn)

            print(u"creating defn-img dict with " + (u"" if defn is None else defn) + u" / " + eng_defn)

            if len(eng_words) != 0:  # don't translate nonexistent words
                translations = {}
                translations["English"] = eng_words

                for lang in multi_defns:
                    words = multi_defns[lang]
                    lang_defn = words[idx]
                    lang_words = self.getDefnWords(lang_defn)
                    translations[lang] = lang_words

                print translations

                bliss_word = Blissymbol(img_filename=imgs[idx],
                                        pos=parts_of_speech[idx],
                                        derivation=derivations[idx],
                                        #language=language,
                                        translations=translations,
                                        translator=self.translator)

                translations = bliss_word.getTranslations()

                for lang in translations:
                    for translation in translations[lang]:
                        lang_bliss_dict.setdefault(translation, [])
                        lang_bliss_dict[translation].append(bliss_word)
                        if lang == "English":
                            bliss_word.addBlissAndUnicode(translation)

            print("\n")
            idx += 1
        raw_input("Press enter to continue.\n")
        return lang_bliss_dict

    # HELPERS
    # =======
    def stripParens(self, word):
        """
        Strips parenthetical(s) from the given word.
        ~
        String cuts off at end of the first predicted lexeme.

        e.g. stripParens("English_(language)") -> "English"

        :param word: str, word to strip parentheticals from
        :return: str, word with parentheticals stripped
        """
        new_word = []
        remove = False

        for char in word:
            if remove == False and char != "(":
                new_word.append(char)
            elif char == "(":
                remove = True
            elif char == ")":
                remove = False

        new_word = ("".join(new_word)).strip()
        return new_word

    def parseAlphabetic(self, word):
        """
        Parses the given non-alphabetic word into an
        alphabetic-only version of the word.

        e.g. parseAlphabetic("English_(language)") -> "English (language)"

        :param word: str, non-alphabetic word
        :return: str, alphabetic version of input word
        """
        word = word.replace("_", " ")
        word = word.replace("-", " ")
        word = word.replace("!", "")
        return word

    def cleanILIDefn(self, defn):
        """
        Cleans the input ILI definition line.

        :param defn: str, line in ILI
        :return: ILIEntry, an entry for a single ILI concept
        """
        idx = re.search(pattern="i[0-9]{1,8}", string=defn).group(0)
        locn = re.search(pattern="[0-9]{8}-[a|n|r|s|v]", string=defn).group(0)
        word = re.search(pattern="#\s.+", string=defn).group(0)
        word = str(word[2:])
        words = word.split(",")
        words = [word.strip() for word in words]
        entry = self.ili_dict.makeEntry(int(idx[1:]), str(locn), words)
        return entry

    def printDict(self, d):
        """
        Prints the given dictionary as if it were written in code.
        Used for visualizing dict contents.

        :param d: dict, input dictionary to print
        :return: None
        """
        print self.dictToStr(d)

    def dictToStr(self, d):
        """
        Returns the given dictionary as the string it would be
        written as in Python.

        :param d: dict, input dictionary to turn to string
        :return: str, input dictionary turned to string
        """
        res = ["{"]
        for key in sorted(d.keys()):
            val = d[key]
            if type(val) == list:
                if len(val) != 0:
                    items = [str(item) for item in val]
                    bliss_names = UNICODE_TO_BLISS[key]
                    comment = "\t# " + ", ".join(bliss_names) + "\n"
                    res.append(comment)
                    res.append('\t"' + str(key) + '":\t[' + ",\n\t\t\t\t".join(items) + '],\n')
                    res.append("\n")
            else:
                res.append('\t"' + str(key) + '": ' + str(d[key]) + ',\n')
        res.append("}")
        return "".join(res)